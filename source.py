# 1.posms Bankas datu analīze S un L uzņēmumā
from openpyxl import load_workbook
wb = load_workbook('S_010123_310523.xlsx')
ws = wb['01-05.2023']
max_row = ws.max_row
summa = 0
for row in ws.iter_rows(min_row=2, max_col=3, max_row=max_row):
    for cell in row:
        if str(cell.value).startswith('B'):
             summa += float(cell.offset(column=3).value)
print(f"S_010123_310523 banka,EUR: {round(summa, 2)}")
wb.close()

wb_L = load_workbook('L_01_05.23.xlsx')
ws_L = wb_L['Banka']
max_row_L = ws_L.max_row
summa_L = 0
for row in ws_L.iter_rows(min_row=3, max_col=3, max_row=max_row_L):
  for cell in row:
    if cell.value is not None:
      if isinstance(cell.value, (int, float)):
        summa_L += cell.value
print(f"L_01_05.23 banka, EUR: {round(summa_L, 2)}")
wb_L.close()
        
# 2.Programmas testēšana uz nelieliem testa failiem
import pandas as pd	
from openpyxl import load_workbook	
from tabulate import tabulate	
	
def replace_minus_with_none(value):	
      return None if pd.isna(value) else float(str(value).replace('-', ''))	

wb_S = load_workbook('S_TESTS.xlsx')	
ws_S = wb_S.active	
max_row_S = ws_S.max_row	
saraksts_S_A = {}	

for row in range(2, max_row_S + 1):	
      A = replace_minus_with_none(ws_S['A' + str(row)].value)	
      saraksts_S_A[A] = replace_minus_with_none(ws_S['B' + str(row)].value)	
	wb_L = load_workbook('L_TESTS.xlsx')	
ws_L = wb_L.active	
max_row_L = ws_L.max_row	
saraksts_L_A = {}	
	
for row in range(1, max_row_L + 1):	
      A = replace_minus_with_none(ws_L['A' + str(row)].value)	
      saraksts_L_A[A] = replace_minus_with_none(ws_L['B' + str(row)].value)	
	
result = []	
	
for key, val_S in saraksts_S_A.items():	
      val_L = saraksts_L_A.get(key)	
      if val_L is not None and val_S != val_L:	
          result.append((key, round(val_S, 2), round(val_L, 2), round(val_S - val_L, 2)))	
	
if result:	
      df = pd.DataFrame(result, columns=['A', 'S_TESTS B', 'L_TESTS B', 'Starpība'])	
      print(tabulate(df, headers='keys', tablefmt='pretty', showindex=False))	
else:	
      print("Nav atšķirību A un B kolonnās.")	
# 3.posms - Visu datu analīze
import pandas as pd	
from openpyxl import load_workbook	
from tabulate import tabulate	

def replace_minus_with_none(value):
  if pd.isna(value) or not (isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('-', '').replace('.', '').isdigit())):
      return None
  try:
      return float(str(value).replace('-', ''))
  except ValueError:
      return None

wb_S = load_workbook('S_010123_310523.xlsx')	
ws_S = wb_S['Sheet7']	
max_row_S = ws_S.max_row	
saraksts_S_A = {}	

for row in range(3, max_row_S-1):	
      A = replace_minus_with_none(ws_S['A' + str(row)].value)	
      saraksts_S_A[A] = replace_minus_with_none(ws_S['B' + str(row)].value)	

wb_L = load_workbook('L_01_05.23.xlsx')	
ws_L = wb_L['Sheet2']		
max_row_L = ws_L.max_row	
saraksts_L_A = {}	

for row in range(1, max_row_L ):	
      A = replace_minus_with_none(ws_L['A' + str(row)].value)	
      saraksts_L_A[A] = replace_minus_with_none(ws_L['B' + str(row)].value)	

result = []	

for key, val_S in saraksts_S_A.items():	
      val_L = saraksts_L_A.get(key)	
      if val_L is not None and val_S != val_L:	
          result.append((key, round(val_S, 2), round(val_L, 2), round(val_S - val_L, 2)))	

if result:	
      df = pd.DataFrame(result, columns=['A', 'S_010123_310523 B', 'L_01_05.23 B', 'Starpība'])	
      print(tabulate(df, headers='keys', tablefmt='pretty', showindex=False))	
else:	
      print("Nav atšķirību A un B kolonnās.")	

